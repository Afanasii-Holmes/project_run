from django.db.models import Sum, Count, Q, Avg, Max
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import viewsets, status
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination

from django_filters.rest_framework import DjangoFilterBackend

from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem, Subscription
from .serializers import RunSerializer, UserSerializer, ChallengeSerializer, PositionSerializer, \
    CollectibleItemSerializer, CoachSerializer, AthleteSerializer
from django.contrib.auth.models import User
from geopy.distance import geodesic
import openpyxl as op


@api_view(['GET'])
def company_details(request):
    print('DEBUG 123456789')
    print('DEBUG FFFFFFFFF')
    return Response({'company_name': 'Бегуны - фантомасы',
                     'slogan':'Бег - это чудо!',
                     'contacts': 'Тел. 222-232-3222'})


class MyPagination(PageNumberPagination):
    page_size_query_param = 'size'


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    pagination_class = MyPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = UserSerializer
    pagination_class = MyPagination
    # queryset = User.objects.filter(is_superuser=False)
    queryset = User.objects.filter(is_superuser=True)

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']

    def get_queryset(self):
        qs = self.queryset
        user_type = self.request.query_params.get('type')
        print('DEBUG user_type', user_type)
        if user_type and user_type=='coach':
            qs = qs.filter(is_staff=True)
        if user_type and user_type=='athlete':
            qs = qs.filter(is_staff=False)
        qs = qs.annotate(runs_finished=Count('run', filter=Q(run__status='finished')))
        qs = qs.annotate(rating=Avg('athletes__rating'))
        print('DEBUG qs', qs)
        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            # return UserDetailSerializer
            user = self.get_object() # крутой метод! почему я его не знал!!
            if user.is_staff:
                return CoachSerializer
            else:
                return AthleteSerializer
        return super().get_serializer_class()


class StatusStartView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == 'init':
            run.status = 'in_progress'
            run.save()
            return Response({'message': 'Все ништяк'}, status=status.HTTP_200_OK)
        else:
            return Response({'message': 'Этот забег стартовать нельзя, он уже стартовал'},
                            status=status.HTTP_400_BAD_REQUEST)


class StatusStopView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == 'in_progress':
            run.status = 'finished'
            if Position.objects.filter(run=run_id).exists():
                # -------------------------------------------
                positions_qs = Position.objects.filter(run=run_id)
                positions_quantity = len(positions_qs)
                distance = 0
                for i in range(positions_quantity-1):
                    distance += geodesic((positions_qs[i].latitude,positions_qs[i].longitude), (positions_qs[i+1].latitude,positions_qs[i+1].longitude)).kilometers
                run.distance = distance
                # -------------------------------------------
                positions_qs_sorted_by_date = positions_qs.order_by('date_time')
                run_time = positions_qs_sorted_by_date[positions_quantity-1].date_time - positions_qs_sorted_by_date[0].date_time
                run.run_time_seconds = run_time.total_seconds()
                #-------------------------------------------
                average_speed = positions_qs.aggregate(Avg('speed'))
                run.speed = round(average_speed['speed__avg'], 2)
                print('DEBUG average_speed', average_speed)
            run.save()
            # -------------------------------------------
            if Run.objects.filter(status='finished', athlete=run.athlete).count() >= 10:
                challenge, created = Challenge.objects.get_or_create(full_name='Сделай 10 Забегов!',
                                                                     athlete=run.athlete)
            # -------------------------------------------
            distance_sum = Run.objects.filter(status='finished', athlete=run.athlete).aggregate(Sum('distance'))

            if distance_sum['distance__sum'] >= 50:
                challenge, created = Challenge.objects.get_or_create(full_name='Пробеги 50 километров!',
                                                                     athlete=run.athlete)
            # -------------------------------------------
            if run.distance >= 2 and run.run_time_seconds <= 600:
                challenge, created = Challenge.objects.get_or_create(
                    full_name='2 километра за 10 минут!',
                    athlete=run.athlete)

            return Response({'message': 'Все ништяк'}, status=status.HTTP_200_OK)
        else:
            return Response({'message': 'Этот забег финишировать нельзя, он еще не стартовал или уже завершен'},
                            status=status.HTTP_400_BAD_REQUEST)


class AthleteInfoView(APIView):

    def get(self, request, user_id):
        user = User.objects.get(id=user_id)
        athlete_info, created = AthleteInfo.objects.get_or_create(user=user)
        return Response({'weight': athlete_info.weight,
                         'goals': athlete_info.goals,
                         'user_id': athlete_info.user.id

        })

    def put(self, request, user_id):
        goals = request.data.get('goals')
        weight = request.data.get('weight')

        if not str(weight).isdigit() or int(weight) <= 0 or int(weight) >= 900:
            return Response({'message': 'weight должен быть числом больше 0 и меньше 900'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(id=user_id).exists():
            user = User.objects.get(id=user_id)
        else:
            return Response({'message':'Такого User не существует'}, status=status.HTTP_404_NOT_FOUND)

        athlete_info, created = AthleteInfo.objects.update_or_create(
            user = user,
            defaults = {'goals':goals, 'weight':weight}
        )

        return Response({'message': 'Создано/изменено'}, status=status.HTTP_201_CREATED)


class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['athlete']


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        qs = self.queryset
        run_id = self.request.query_params.get('run')
        if run_id:
            qs = qs.filter(run=run_id)
        return qs

    def perform_create(self, serializer):
        run = serializer.validated_data['run']
        serializer.save()  # чтобы включить только что создаваемую позицию в QS
        all_positions = Position.objects.filter(run=run)
        if all_positions.count() > 1:
            ordered_positions = all_positions.order_by('-id')
            last_position = ordered_positions[0]
            previous_position = ordered_positions[1]
            previous_distance = previous_position.distance
            last_distance = geodesic((last_position.latitude, last_position.longitude),
                                     (previous_position.latitude, previous_position.longitude)).meters
            time_delta = last_position.date_time - previous_position.date_time
            speed = last_distance / time_delta.total_seconds()
            last_position.speed = round(speed, 2)
            last_position.distance = round(previous_distance + last_distance / 1000, 2)
            last_position.save()

    # def create(self, request, pk=None): # Добавил 28 мая, версия студента, можно удалить
    #     data = request.data
    #     run_id = data.get("run", None)
    #     try:
    #         run = Run.objects.get(id=run_id)
    #     except Run.DoesNotExist:
    #         return Response(
    #             {"detail": "Забег не найден"}, status.HTTP_400_BAD_REQUEST)
    #     if run.status != "in_progress":
    #         return Response(
    #             {"detail": "Забег должен быть в статусе 'in_progress'"}, status.HTTP_400_BAD_REQUEST
    #         )
    #     serializer = PositionSerializer(run,data=data,partial=True)
    #     # serializer = PositionSerializer(data=data)
    #
    #     if serializer.is_valid():
    #         serializer.save()
    #         super().create(request, pk)
    #         return Response(serializer.data, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # def create(self, request, pk=None): # Claude version - работает!
    #     data = request.data
    #     run_id = data.get("run")
    #
    #     if not run_id:
    #         return Response(
    #             {"detail": "Поле 'run' обязательно"},
    #             status=status.HTTP_400_BAD_REQUEST
    #         )
    #
    #     try:
    #         run = Run.objects.get(id=run_id)
    #     except Run.DoesNotExist:
    #         return Response(
    #             {"detail": "Забег не найден"},
    #             status=status.HTTP_400_BAD_REQUEST
    #         )
    #
    #     if run.status != "in_progress":
    #         return Response(
    #             {"detail": "Забег должен быть в статусе 'in_progress'"},
    #             status=status.HTTP_400_BAD_REQUEST
    #         )
    #
    #     # Создаем Position, а не обновляем Run
    #     serializer = self.get_serializer(data=data)
    #     if serializer.is_valid():
    #         serializer.save()
    #         return Response(serializer.data, status=status.HTTP_201_CREATED)
    #
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


@api_view(['POST'])
def upload_view(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_xlsx_file = request.FILES['file']
        wb = op.load_workbook(uploaded_xlsx_file, data_only=True)
        sheet = wb.active
        wrong_rows_list = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            name, uid, value, latitude, longitude, picture = row
            data = {
                'name': name,
                'uid': uid,
                'latitude': latitude,
                'longitude': longitude,
                'picture': picture,
                'value': value,
            }
            serializer = CollectibleItemSerializer(data=data)
            if serializer.is_valid():
                CollectibleItem.objects.create(name=name,
                                               uid=uid,
                                               value=value,
                                               latitude=latitude,
                                               longitude=longitude,
                                               picture=picture)
            else:
                wrong_rows_list.append([name, uid, value, latitude, longitude, picture])

        return Response(wrong_rows_list)
    return Response([])


class SubscribeView(APIView):
    def post(self, request, id):
        coach_id = id
        athlete_id = self.request.data['athlete']

        coach = get_object_or_404(User, id=coach_id)
        if not User.objects.filter(id=athlete_id).exists():
            return Response({'message': f'Пользователя c id {athlete_id} не существует'},
                            status=status.HTTP_400_BAD_REQUEST)
        athlete=User.objects.get(id=athlete_id)

        if not coach.is_staff:
            return Response({'message': f'Пользователь c id {coach_id} это не тренер'}, status=status.HTTP_400_BAD_REQUEST)
        if athlete.is_staff:
            return Response({'message': f'Пользователь c id {coach_id} это не бегун'}, status=status.HTTP_400_BAD_REQUEST)
        if Subscription.objects.filter(coach=coach, athlete=athlete).exists():
            return Response({'message': 'Такая подписка уже существует'},
                            status=status.HTTP_400_BAD_REQUEST)
        Subscription.objects.create(coach=coach, athlete=athlete)

        return Response({'message': 'Все ништяк'}, status=status.HTTP_200_OK)


@api_view(['GET'])
def challenge_summary_view(request):
    # Получим список уникальных названий челленджей
    unique_challenges = Challenge.objects.values('full_name').distinct()
    unique_name_list = [i['full_name'] for i in unique_challenges]

    #Получим список атлетов для каждого челленджа
    final_list = []
    for challenge_name in unique_name_list:
        user_queryset = User.objects.filter(challenge__full_name=challenge_name)
        athletes_list = []
        for athlete in user_queryset:
            athletes_list.append({'id': athlete.id, 'full_name': f'{athlete.first_name} {athlete.last_name}',
                                  'username': athlete.username})

        final_list.append({'name_to_display':challenge_name, 'athletes':athletes_list})

    return Response(final_list)


class CoachRatingView(APIView):
    def post(self, request, coach_id):
        athlete_id = request.data.get('athlete')
        rating = request.data.get('rating')

        user_coach = get_object_or_404(User, id=coach_id)

        if not athlete_id:
            return Response({'message': f'Нет поля athlete_id'}, status=status.HTTP_400_BAD_REQUEST)

        if not rating:
            return Response({'message': f'Нет поля rating'}, status=status.HTTP_400_BAD_REQUEST)

        if isinstance(rating, str) and not rating.isdigit():
            return Response({'message': f'У вас rating не цифра. Ваше значение {rating}'},
                            status=status.HTTP_400_BAD_REQUEST)

        if not 1<= int(rating) <=5:
            return Response({'message': f'rating не в пределах от 1 до 5. Ваше значение {rating}'}, status=status.HTTP_400_BAD_REQUEST)

        user_athlete = get_object_or_404(User, id=athlete_id)

        if Subscription.objects.filter(coach=coach_id, athlete=athlete_id).exists():
            subscription = Subscription.objects.get(coach=coach_id, athlete=athlete_id)
            subscription.rating = rating
            subscription.save()
        else:
            return Response({'message':f'Бегун c id {athlete_id} не подписан на тренера с id {coach_id}'},
                            status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': 'ОК'})


class AnalyticsCoachView(APIView):
    def get(self, request, coach_id):
        coach_queryset = Subscription.objects.filter(coach=coach_id)

        qs_with_additional_fields = coach_queryset.annotate(
            max_distance=Max('athlete__run__distance'),
            sum_distances=Sum('athlete__run__distance'),
            avg_speed=Avg('athlete__run__speed')
        )

        longest_qs = qs_with_additional_fields.order_by('-max_distance').first()
        longest_run_value = longest_qs.max_distance
        longest_run_user = longest_qs.athlete_id

        max_total_run_qs = qs_with_additional_fields.order_by('-sum_distances').first()
        total_run_value = max_total_run_qs.sum_distances
        total_run_user = max_total_run_qs.athlete_id

        max_avg_speed_qs = qs_with_additional_fields.order_by('-avg_speed').first()
        speed_avg_value = max_avg_speed_qs.avg_speed
        speed_avg_user = max_avg_speed_qs.athlete_id

        return Response({'longest_run_value': longest_run_value,
                         'longest_run_user': longest_run_user,
                         'total_run_value': total_run_value,
                         'total_run_user': total_run_user,
                         'speed_avg_value': speed_avg_value,
                         'speed_avg_user': speed_avg_user
                         })